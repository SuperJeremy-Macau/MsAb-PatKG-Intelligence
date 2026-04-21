from __future__ import annotations

import argparse
import json
import os
import re
import sys
from pathlib import Path
from typing import Dict, List

import openpyxl

ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
if ROOT not in sys.path:
    sys.path.insert(0, ROOT)


def _norm_exact(value: str) -> str:
    return " ".join(str(value).strip().upper().split())


def _norm_components(value: str) -> str:
    parts = [p.strip().upper() for p in str(value).split("/") if p and p.strip()]
    return "/".join(sorted(parts))


def _norm_loose(value: str) -> str:
    value = str(value).strip().upper()
    return re.sub(r"[^A-Z0-9/\-\.]+", "", value)


def _load_excel_targetpairs(excel_path: Path) -> List[str]:
    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise RuntimeError(f"Workbook is empty: {excel_path}")

    header = [str(x).strip() if x is not None else "" for x in rows[0]]
    if "TargetPair" not in header:
        raise RuntimeError(f"'TargetPair' column not found in {excel_path}")
    idx = header.index("TargetPair")
    return [
        str(row[idx]).strip()
        for row in rows[1:]
        if row[idx] is not None and str(row[idx]).strip()
    ]


def _load_neo4j_targetpairs(settings_path: str) -> List[str]:
    from bsab_kg_qa_en.config import load_settings
    from bsab_kg_qa_en.kg import Neo4jRunner

    cfg = load_settings(settings_path)
    neo = cfg["neo4j"]
    runner = Neo4jRunner(
        uri=neo["uri"],
        user=neo["user"],
        password=neo["password"],
        database=neo["database"],
        max_rows=int(neo.get("max_rows", 50)),
    )
    try:
        rows = runner.run(
            """
            MATCH (tp:TargetPair)
            WHERE tp.name IS NOT NULL
            RETURN DISTINCT toString(tp.name) AS name
            ORDER BY name
            """,
            enforce_limit=False,
        )
    finally:
        runner.close()
    return [str(row["name"]).strip() for row in rows if row.get("name")]


def _group_by(values: List[str], normalizer) -> Dict[str, List[str]]:
    out: Dict[str, List[str]] = {}
    for value in values:
        out.setdefault(normalizer(value), []).append(value)
    return out


def _build_report(excel_values: List[str], neo_values: List[str]) -> dict:
    excel_unique = sorted(set(excel_values))
    neo_unique = sorted(set(neo_values))

    excel_exact_map = _group_by(excel_unique, _norm_exact)
    neo_exact_map = _group_by(neo_unique, _norm_exact)
    excel_loose_map = _group_by(excel_unique, _norm_loose)
    neo_loose_map = _group_by(neo_unique, _norm_loose)
    neo_component_map = _group_by(neo_unique, _norm_components)

    exact_missing_in_neo = sorted(
        value for value in excel_unique if _norm_exact(value) not in neo_exact_map
    )
    exact_extra_in_neo = sorted(
        value for value in neo_unique if _norm_exact(value) not in excel_exact_map
    )

    symbol_or_encoding_mismatches = []
    for key in sorted(set(excel_loose_map) & set(neo_loose_map)):
        excel_group = sorted(excel_loose_map[key])
        neo_group = sorted(neo_loose_map[key])
        if excel_group != neo_group:
            symbol_or_encoding_mismatches.append(
                {"normalized": key, "excel": excel_group, "neo4j": neo_group}
            )

    component_reorder_candidates = []
    for value in exact_missing_in_neo:
        component_key = _norm_components(value)
        if component_key in neo_component_map:
            component_reorder_candidates.append(
                {"excel": value, "neo4j_candidates": sorted(neo_component_map[component_key])}
            )

    true_missing_in_neo = sorted(
        value for value in excel_unique if _norm_loose(value) not in neo_loose_map
    )
    true_extra_in_neo = sorted(
        value for value in neo_unique if _norm_loose(value) not in excel_loose_map
    )

    singletons = sorted(value for value in neo_unique if "/" not in value)
    self_pairs = []
    for value in neo_unique:
        parts = [p.strip() for p in value.split("/") if p.strip()]
        if len(parts) >= 2 and len(set(parts)) == 1:
            self_pairs.append(value)
    malformed_encoding = sorted(value for value in neo_unique if "\ufffd" in value)
    triple_or_more = sorted(value for value in neo_unique if value.count("/") >= 2)

    return {
        "excel": {
            "row_count": len(excel_values),
            "unique_targetpair_count": len(excel_unique),
        },
        "neo4j": {
            "unique_targetpair_count": len(neo_unique),
        },
        "exact_comparison": {
            "missing_in_neo_count": len(exact_missing_in_neo),
            "extra_in_neo_count": len(exact_extra_in_neo),
            "missing_in_neo": exact_missing_in_neo,
            "extra_in_neo": exact_extra_in_neo,
            "component_reorder_candidates": component_reorder_candidates,
        },
        "loose_normalization_comparison": {
            "likely_symbol_or_encoding_mismatches": symbol_or_encoding_mismatches,
            "true_missing_in_neo_count": len(true_missing_in_neo),
            "true_missing_in_neo": true_missing_in_neo,
            "true_extra_in_neo_count": len(true_extra_in_neo),
            "true_extra_in_neo": true_extra_in_neo,
        },
        "neo4j_anomalies": {
            "singleton_count": len(singletons),
            "singletons": singletons,
            "self_pair_count": len(self_pairs),
            "self_pairs": sorted(self_pairs),
            "replacement_char_count": len(malformed_encoding),
            "replacement_char_values": malformed_encoding,
            "triple_or_more_count": len(triple_or_more),
            "triple_or_more_examples": triple_or_more[:100],
        },
    }


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Compare Excel TargetPair-TechnologyClass mapping against current Neo4j TargetPair nodes."
    )
    parser.add_argument("--excel", default="docs/TargetPair-TechnologyClass.xlsx")
    parser.add_argument("--settings", default="bsab_kg_qa_en/config/settings.yaml")
    parser.add_argument("--output", default="")
    args = parser.parse_args()

    excel_values = _load_excel_targetpairs(Path(args.excel))
    neo_values = _load_neo4j_targetpairs(args.settings)
    report = _build_report(excel_values, neo_values)

    if args.output:
        output_path = Path(args.output)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        output_path.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")

    summary = {
        "excel_unique_targetpairs": report["excel"]["unique_targetpair_count"],
        "neo4j_unique_targetpairs": report["neo4j"]["unique_targetpair_count"],
        "true_missing_in_neo_count": report["loose_normalization_comparison"]["true_missing_in_neo_count"],
        "true_extra_in_neo_count": report["loose_normalization_comparison"]["true_extra_in_neo_count"],
        "likely_symbol_or_encoding_mismatch_count": len(
            report["loose_normalization_comparison"]["likely_symbol_or_encoding_mismatches"]
        ),
        "neo4j_singleton_count": report["neo4j_anomalies"]["singleton_count"],
        "neo4j_self_pair_count": report["neo4j_anomalies"]["self_pair_count"],
        "neo4j_replacement_char_count": report["neo4j_anomalies"]["replacement_char_count"],
    }
    print(json.dumps(summary, ensure_ascii=True, indent=2))
    if args.output:
        print(f"Full report written to: {args.output}")


if __name__ == "__main__":
    main()
